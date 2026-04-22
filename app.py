"""
app.py — BioField Dashboard  (v4 — Cloud Edition)
author: Dr. Hakim Mitiche  |  Flask UI by Claude

Storage:
  - Images  → Cloudinary (object storage, permanent URLs)
  - Data    → PostgreSQL  (all observations, metadata, processing state)

No local files required. Works on Render, Railway, Fly.io, etc.
"""

from __future__ import annotations

import csv
import io
import json
import logging
import os
import queue
import subprocess
import sys
import threading
import zipfile
from collections import Counter, defaultdict
from pathlib import Path

import cloudinary
import cloudinary.uploader
import psycopg2
import psycopg2.extras
from dotenv import load_dotenv
from flask import (Flask, Response, jsonify, render_template,
                   request, send_file, stream_with_context)

load_dotenv()

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
)

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024   # 50 MB per upload

# ── Config ─────────────────────────────────────────────────────────────────────

BASE_DIR = Path(__file__).parent

CATEGORIES = {
    "insect": {"label": "Insects", "color": "#f59e0b", "icon": "🪲"},
    "flora":  {"label": "Flora",   "color": "#4ade80", "icon": "🌿"},
    "fungus": {"label": "Fungi",   "color": "#c084fc", "icon": "🍄"},
}

ALLOWED_EXT = {".jpg", ".jpeg", ".png", ".webp"}

# ── Cloudinary ─────────────────────────────────────────────────────────────────
# Use .get() so the app doesn't crash at startup when .env is missing.
# A helpful error is shown at runtime instead of a silent 404 for all routes.

cloudinary.config(
    cloud_name = os.environ.get("CLOUDINARY_CLOUD_NAME", ""),
    api_key    = os.environ.get("CLOUDINARY_API_KEY",    ""),
    api_secret = os.environ.get("CLOUDINARY_API_SECRET", ""),
    secure     = True,
)

# ── PostgreSQL ─────────────────────────────────────────────────────────────────

def get_db():
    """Return a new psycopg2 connection."""
    db_url = os.environ.get("DATABASE_URL", "")
    if not db_url:
        raise RuntimeError(
            "DATABASE_URL is not set. Add it to your .env file "
            "or Render environment variables."
        )
    return psycopg2.connect(db_url, sslmode="require")


def init_db():
    """Create tables if they don't exist yet. Called once at startup."""
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                CREATE TABLE IF NOT EXISTS observations (
                    id                SERIAL PRIMARY KEY,
                    category          VARCHAR(20)  NOT NULL,
                    picture_name      VARCHAR(255),
                    cloudinary_url    TEXT,
                    cloudinary_public_id VARCHAR(255),
                    common_name       VARCHAR(255),
                    scientific_name   VARCHAR(255),
                    species_name      VARCHAR(255),
                    date              VARCHAR(50),
                    gps_string        TEXT,
                    latitude_dd       DOUBLE PRECISION,
                    longitude_dd      DOUBLE PRECISION,
                    altitude_m        DOUBLE PRECISION,
                    processing_status VARCHAR(50)  DEFAULT 'PENDING',
                    created_at        TIMESTAMP    DEFAULT NOW()
                );

                CREATE INDEX IF NOT EXISTS idx_obs_category
                    ON observations(category);
                CREATE INDEX IF NOT EXISTS idx_obs_status
                    ON observations(processing_status);
            """)
        conn.commit()
    app.logger.info("Database initialised.")


# ── Data Helpers ───────────────────────────────────────────────────────────────

def get_all_observations(category: str = "all") -> list[dict]:
    with get_db() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            if category == "all":
                cur.execute("SELECT * FROM observations ORDER BY created_at DESC")
            else:
                cur.execute(
                    "SELECT * FROM observations WHERE category = %s ORDER BY created_at DESC",
                    (category,)
                )
            rows = cur.fetchall()
    # Convert RealDictRow → plain dict, add _category alias for templates
    result = []
    for r in rows:
        d = dict(r)
        d["_category"] = d["category"]
        result.append(d)
    return result


def get_stats() -> dict:
    stats = {"total": 0, "with_gps": 0, "categories": {}, "db_processed": 0}

    with get_db() as conn:
        with conn.cursor() as cur:
            # Per-category totals
            cur.execute("""
                SELECT category,
                       COUNT(*)                                          AS total,
                       COUNT(*) FILTER (WHERE gps_string IS NOT NULL
                                          AND gps_string != 'No GPS Data'
                                          AND gps_string != '')          AS with_gps,
                       COUNT(*) FILTER (WHERE cloudinary_url IS NOT NULL) AS images_in_cloud
                FROM observations
                GROUP BY category
            """)
            for cat, total, gps, imgs in cur.fetchall():
                stats["categories"][cat] = {
                    "total":          total,
                    "with_gps":       gps,
                    "images_on_disk": imgs,   # reuse field name so templates work unchanged
                    "label": CATEGORIES.get(cat, {}).get("label", cat),
                    "color": CATEGORIES.get(cat, {}).get("color", "#888"),
                    "icon":  CATEGORIES.get(cat, {}).get("icon",  "?"),
                }
                stats["total"]    += total
                stats["with_gps"] += gps

            # Ensure all categories exist even if empty
            for cat, cfg in CATEGORIES.items():
                if cat not in stats["categories"]:
                    stats["categories"][cat] = {
                        "total": 0, "with_gps": 0, "images_on_disk": 0,
                        "label": cfg["label"], "color": cfg["color"], "icon": cfg["icon"],
                    }

            # Processed count
            cur.execute("SELECT COUNT(*) FROM observations WHERE processing_status = 'SUCCESS'")
            stats["db_processed"] = cur.fetchone()[0]

    # Map file no longer stored locally on Render; mark as not available by default.
    # (map.py can upload the generated HTML to Cloudinary or store as a DB record too)
    stats["map_exists"]     = (BASE_DIR / "bio_observations_map.html").exists()
    stats["geojson_exists"] = (BASE_DIR / "bio_observations.geojson").exists()
    return stats


def get_recent_observations(limit: int = 10) -> list[dict]:
    with get_db() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                "SELECT * FROM observations ORDER BY created_at DESC LIMIT %s",
                (limit,)
            )
            rows = cur.fetchall()
    result = []
    for r in rows:
        d = dict(r)
        d["_category"] = d["category"]
        result.append(d)
    return result


def get_analytics_data() -> dict:
    rows = get_all_observations()

    month_counts: dict[str, int] = defaultdict(int)
    for r in rows:
        d = r.get("date", "") or ""
        if d and d != "N/A" and len(d) >= 7:
            month_counts[d[:7]] += 1
    months_sorted = sorted(month_counts.items())

    species_by_cat: dict[str, set] = defaultdict(set)
    for r in rows:
        name = (r.get("common_name") or r.get("species_name") or "").strip()
        if name and name.lower() not in ("unknown", ""):
            species_by_cat[r["_category"]].add(name)

    status_counts: Counter = Counter()
    for r in rows:
        st = r.get("processing_status") or "UNKNOWN"
        status_counts[st] += 1

    alt_buckets = {"0–500 m": 0, "500–1000 m": 0, "1000–1500 m": 0, "1500+ m": 0}
    for r in rows:
        try:
            a = float(r.get("altitude_m") or 0)
            if a < 500:       alt_buckets["0–500 m"]    += 1
            elif a < 1000:    alt_buckets["500–1000 m"] += 1
            elif a < 1500:    alt_buckets["1000–1500 m"]+= 1
            else:             alt_buckets["1500+ m"]    += 1
        except (ValueError, TypeError):
            pass

    all_species: Counter = Counter()
    for r in rows:
        name = (r.get("common_name") or r.get("species_name") or "").strip()
        if name and name.lower() not in ("unknown", ""):
            all_species[name] += 1

    with_gps    = sum(1 for r in rows
                      if (r.get("gps_string") or "").strip() not in ("", "No GPS Data"))
    without_gps = len(rows) - with_gps

    return {
        "total": len(rows),
        "timeline": {
            "labels": [m[0] for m in months_sorted],
            "values": [m[1] for m in months_sorted],
        },
        "species_richness": {cat: len(sp) for cat, sp in species_by_cat.items()},
        "status":           dict(status_counts),
        "altitude":         alt_buckets,
        "top_species":      all_species.most_common(10),
        "gps_coverage":     {"With GPS": with_gps, "No GPS": without_gps},
    }


# ── SSE Streaming ──────────────────────────────────────────────────────────────

_log_queues: dict[str, queue.Queue] = {}


def _stream_subprocess(cmd: list[str], stream_id: str):
    q = _log_queues[stream_id]
    try:
        env = {**os.environ}   # pass all env vars (DATABASE_URL, CLOUDINARY_*, etc.)
        proc = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                                text=True, bufsize=1, cwd=str(BASE_DIR), env=env)
        for line in proc.stdout:
            q.put({"type": "log", "text": line.rstrip()})
        proc.wait()
        q.put({"type": "done", "rc": proc.returncode,
               "text": f"[Exit code {proc.returncode}]"})
    except Exception as e:
        q.put({"type": "error", "text": str(e)})
    finally:
        q.put(None)


def sse_generator(stream_id: str):
    q = _log_queues.get(stream_id)
    if not q:
        yield "data: {}\n\n"
        return
    while True:
        item = q.get()
        if item is None:
            yield "data: " + json.dumps({"type": "end"}) + "\n\n"
            _log_queues.pop(stream_id, None)
            break
        yield "data: " + json.dumps(item) + "\n\n"


# ── Global context ─────────────────────────────────────────────────────────────

@app.context_processor
def inject_globals():
    return {"categories": CATEGORIES}


# ── Page Routes ────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html", stats=get_stats(), recent=get_recent_observations(8))


@app.route("/observations")
def observations():
    cat_filter = request.args.get("category", "all")
    all_obs    = get_all_observations(cat_filter)
    return render_template("observations.html", observations=all_obs,
                           active=cat_filter, total=len(all_obs))


@app.route("/map")
def map_view():
    map_file = BASE_DIR / "bio_observations_map.html"
    geo_file = BASE_DIR / "bio_observations.geojson"
    return render_template("map_view.html",
                           map_exists=map_file.exists(),
                           geojson_exists=geo_file.exists())


@app.route("/map-content")
def map_content():
    map_file = BASE_DIR / "bio_observations_map.html"
    if not map_file.exists():
        return "Map not generated yet.", 404
    return map_file.read_text(encoding="utf-8"), 200, {"Content-Type": "text/html"}


@app.route("/analytics")
def analytics():
    return render_template("analytics.html", data=get_analytics_data())


@app.route("/upload")
def upload_page():
    stats = get_stats()
    cats_with_counts = {}
    for cat, cfg in CATEGORIES.items():
        cats_with_counts[cat] = dict(cfg)
        cats_with_counts[cat]["images_on_disk"] = stats["categories"].get(
            cat, {}).get("images_on_disk", 0)
    return render_template("upload.html", categories=cats_with_counts)


@app.route("/pipeline")
def pipeline():
    api_key_set = bool(os.environ.get("OPENROUTER_API_KEY", "").strip())
    return render_template("pipeline.html", stats=get_stats(), api_key_set=api_key_set)


@app.route("/manage")
def manage():
    """Database management panel."""
    return render_template("manage.html")


# ── Data API ───────────────────────────────────────────────────────────────────

@app.route("/api/stats")
def api_stats():
    return jsonify(get_stats())


@app.route("/api/analytics")
def api_analytics():
    return jsonify(get_analytics_data())

@app.route("/admin")
def admin():
    key = request.args.get("key")

    if key != os.environ.get("ADMIN_KEY", "1234"):
        return "Unauthorized", 403

    stats = get_stats()
    recent = get_recent_observations(20)

    return render_template("admin.html", stats=stats, observations=recent)

# ── Export API ─────────────────────────────────────────────────────────────────
#an API for delete 
@app.route("/api/delete/<int:id>", methods=["DELETE"])
def delete_item(id):
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM observations WHERE id = %s", (id,))
        conn.commit()
    return jsonify({"status": "deleted"})

@app.route("/api/geojson")
def api_geojson():
    rows = get_all_observations()
    features = []
    for r in rows:
        try:
            lat = float(r.get("latitude_dd") or 0)
            lon = float(r.get("longitude_dd") or 0)
            if lat == 0 and lon == 0:
                continue
        except (ValueError, TypeError):
            continue
        features.append({
            "type": "Feature",
            "geometry": {"type": "Point", "coordinates": [lon, lat]},
            "properties": {
                "picture_name":    r.get("picture_name", ""),
                "category":        r["_category"],
                "species_name":    r.get("common_name") or r.get("species_name", ""),
                "scientific_name": r.get("scientific_name", ""),
                "date":            r.get("date", ""),
                "altitude":        r.get("altitude_m", ""),
                "cloudinary_url":  r.get("cloudinary_url", ""),
            },
        })
    geojson_str = json.dumps({"type": "FeatureCollection", "features": features},
                             ensure_ascii=False, indent=2)
    return Response(geojson_str, mimetype="application/geo+json",
                    headers={"Content-Disposition":
                             "attachment; filename=bio_observations.geojson"})


@app.route("/api/export-csv")
def api_export_csv():
    cat_filter = request.args.get("category", "all")
    rows = get_all_observations(cat_filter)
    if not rows:
        return jsonify({"error": "No data"}), 404

    # Exclude internal fields
    skip = {"_category", "id"}
    fieldnames = [k for k in rows[0].keys() if k not in skip]

    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(rows)

    fname = f"bio_{cat_filter}.csv"
    return Response(buf.getvalue(), mimetype="text/csv",
                    headers={"Content-Disposition": f"attachment; filename={fname}"})


@app.route("/api/export-excel")
def api_export_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({"error": "openpyxl not installed."}), 500

    cat_filter = request.args.get("category", "all")
    rows = get_all_observations(cat_filter)
    if not rows:
        return jsonify({"error": "No data"}), 404

    skip = {"_category", "id"}
    wb = openpyxl.Workbook()
    sheets_data = {}
    for r in rows:
        sheets_data.setdefault(r["_category"], []).append(r)

    CAT_COLORS = {"insect": "FFF59E0B", "flora": "FF4ADE80", "fungus": "FFC084FC"}

    for idx, (cat, cat_rows) in enumerate(sheets_data.items()):
        ws = wb.active if idx == 0 else wb.create_sheet()
        ws.title = cat.capitalize()
        fieldnames = [k for k in cat_rows[0].keys() if k not in skip]
        hfill = PatternFill("solid", fgColor=CAT_COLORS.get(cat, "FF4ADE80"))
        hfont = Font(bold=True, color="FF000000")
        for col_i, field in enumerate(fieldnames, 1):
            cell = ws.cell(row=1, column=col_i, value=field.replace("_", " ").title())
            cell.fill = hfill; cell.font = hfont
            cell.alignment = Alignment(horizontal="center")
        for row_i, r in enumerate(cat_rows, 2):
            for col_i, field in enumerate(fieldnames, 1):
                ws.cell(row=row_i, column=col_i, value=r.get(field, ""))
        for col_i in range(1, len(fieldnames) + 1):
            ws.column_dimensions[get_column_letter(col_i)].auto_size = True

    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f"bio_{cat_filter}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/api/backup")
def api_backup():
    """Export all observations as a CSV + GeoJSON ZIP (no local DB to back up)."""
    rows = get_all_observations()
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        # CSV per category
        for cat in CATEGORIES:
            cat_rows = [r for r in rows if r["_category"] == cat]
            if cat_rows:
                skip = {"_category", "id"}
                fieldnames = [k for k in cat_rows[0].keys() if k not in skip]
                sbuf = io.StringIO()
                writer = csv.DictWriter(sbuf, fieldnames=fieldnames, extrasaction="ignore")
                writer.writeheader(); writer.writerows(cat_rows)
                zf.writestr(f"bio_{cat}.csv", sbuf.getvalue())
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="biofield_backup.zip",
                     mimetype="application/zip")


# ── Image API — now just redirects to Cloudinary URL ──────────────────────────

@app.route("/api/image-url/<int:obs_id>")
def api_image_url(obs_id: int):
    """Return the Cloudinary URL for a given observation id."""
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT cloudinary_url FROM observations WHERE id = %s", (obs_id,))
            row = cur.fetchone()
    if not row or not row[0]:
        return "Not found", 404
    return jsonify({"url": row[0]})



# ── DB Management API ──────────────────────────────────────────────────────────

@app.route("/api/db-stats")
def api_db_stats():
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) FROM observations")
            total = cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM observations WHERE processing_status='SUCCESS'")
            success = cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM observations WHERE processing_status='PENDING'")
            pending = cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM observations WHERE processing_status='AI_FAILED'")
            failed = cur.fetchone()[0]
    return jsonify({"total": total, "success": success, "pending": pending, "failed": failed})


@app.route("/api/db-records")
def api_db_records():
    with get_db() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("""
                SELECT id, category, picture_name, cloudinary_url,
                       common_name, species_name, date, processing_status
                FROM observations ORDER BY id DESC
            """)
            rows = [dict(r) for r in cur.fetchall()]
    return jsonify({"records": rows})


@app.route("/api/db-delete/<int:obs_id>", methods=["DELETE"])
def api_db_delete_one(obs_id: int):
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM observations WHERE id = %s", (obs_id,))
            deleted = cur.rowcount
        conn.commit()
    if deleted:
        return jsonify({"message": f"Record #{obs_id} deleted."})
    return jsonify({"message": "Record not found."}), 404


@app.route("/api/db-delete-status/<status>", methods=["POST"])
def api_db_delete_by_status(status: str):
    allowed = {"AI_FAILED", "PENDING", "DOWNLOAD_ERROR", "NO_URL", "EXIF_ERROR"}
    if status not in allowed:
        return jsonify({"message": f"Status not allowed."}), 400
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM observations WHERE processing_status = %s", (status,))
            deleted = cur.rowcount
        conn.commit()
    return jsonify({"message": f"Deleted {deleted} records with status {status}."})


@app.route("/api/db-retry-failed", methods=["POST"])
def api_db_retry_failed():
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                UPDATE observations
                SET processing_status='PENDING', common_name=NULL,
                    scientific_name=NULL, species_name=NULL
                WHERE processing_status='AI_FAILED'
            """)
            updated = cur.rowcount
        conn.commit()
    return jsonify({"message": f"Reset {updated} AI_FAILED records to PENDING."})


@app.route("/api/db-reset", methods=["POST"])
def api_db_reset():
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM observations")
            deleted = cur.rowcount
        conn.commit()
    return jsonify({"message": f"Database reset — {deleted} records deleted."})


# ── Upload API — saves to Cloudinary, inserts row in PostgreSQL ────────────────

@app.route("/api/upload", methods=["POST"])
def api_upload():
    app.logger.info("Upload request received")
    category = request.form.get("category", "unsorted")
    files    = request.files.getlist("images")
    app.logger.info(f"Category: {category} | Files: {len(files)}")

    if category not in CATEGORIES:
        return jsonify({"status": "error", "message": f"Unknown category: {category}"}), 400
    if not files:
        return jsonify({"status": "error", "message": "No files received"}), 400

    saved_urls = []
    skipped    = []

    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                for f in files:
                    if not f or not f.filename:
                        continue
                    ext = "." + f.filename.rsplit(".", 1)[-1].lower() if "." in f.filename else ""
                    if ext not in ALLOWED_EXT:
                        skipped.append(f.filename)
                        continue

                    # Upload to Cloudinary
                    result = cloudinary.uploader.upload(
                        f,
                        folder=f"biofield/{category}",
                        resource_type="image",
                    )
                    url       = result["secure_url"]
                    public_id = result["public_id"]
                    app.logger.info(f"Cloudinary OK: {url}")

                    # Insert a PENDING row in PostgreSQL
                    cur.execute("""
                        INSERT INTO observations
                            (category, picture_name, cloudinary_url,
                             cloudinary_public_id, processing_status)
                        VALUES (%s, %s, %s, %s, 'PENDING')
                        RETURNING id
                    """, (category, f.filename, url, public_id))
                    obs_id = cur.fetchone()[0]

                    saved_urls.append({"url": url, "id": obs_id, "name": f.filename})

            conn.commit()

        return jsonify({
            "status":      "success",
            "total_saved": len(saved_urls),
            "urls":        saved_urls,
            "skipped":     skipped,
            "message":     "Images uploaded to Cloudinary and registered in database.",
        })

    except Exception as e:
        app.logger.error(f"Upload failed: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


# ── Pipeline API ───────────────────────────────────────────────────────────────

@app.route("/api/run-extract", methods=["POST"])
def api_run_extract():
    import uuid
    sid = str(uuid.uuid4())
    _log_queues[sid] = queue.Queue()
    threading.Thread(target=_stream_subprocess,
                     args=([sys.executable, str(BASE_DIR / "extract.py")], sid),
                     daemon=True).start()
    return jsonify({"stream_id": sid})


@app.route("/api/run-map", methods=["POST"])
def api_run_map():
    import uuid
    sid = str(uuid.uuid4())
    _log_queues[sid] = queue.Queue()
    threading.Thread(target=_stream_subprocess,
                     args=([sys.executable, str(BASE_DIR / "map.py")], sid),
                     daemon=True).start()
    return jsonify({"stream_id": sid})


@app.route("/api/logs/<stream_id>")
def api_logs(stream_id: str):
    return Response(stream_with_context(sse_generator(stream_id)),
                    mimetype="text/event-stream",
                    headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


# ── Startup ────────────────────────────────────────────────────────────────────

with app.app_context():
    try:
        init_db()
    except Exception as e:
        app.logger.warning(f"DB init skipped (no DATABASE_URL?): {e}")

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5050))
    app.run(host="0.0.0.0", port=port, threaded=True)
